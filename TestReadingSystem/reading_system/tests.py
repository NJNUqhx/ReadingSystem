import copy

from reading_system import models
from pypinyin import pinyin, Style
import openpyxl
import time


def GetPinyin(ch):
    res = pinyin(ch, heteronym=True)
    return res[0]


def PinyinToStr(ch):
    res = GetPinyin(ch)
    s = ""
    for i in res:
        s = s + i + ";"
    return s


def GetPinyin2(ch):
    sheng_diao = pinyin(ch, style=Style(0), heteronym=True)[0]
    return sheng_diao


def CompareChar(ch1, ch2):
    py1 = GetPinyin(ch1)
    py2 = GetPinyin(ch2)
    for i in py1:
        if i in py2:
            return True
    return False


def IsSimilarChar(ch1, ch2):
    sheng_diao1 = pinyin(ch1, style=Style(0), heteronym=True)[0]
    sheng_mu1 = pinyin(ch1, style=Style(3), heteronym=True)[0]
    yun_mu1 = pinyin(ch1, style=Style(5), heteronym=True)[0]

    sheng_diao2 = pinyin(ch2, style=Style(0), heteronym=True)[0]
    sheng_mu2 = pinyin(ch2, style=Style(3), heteronym=True)[0]
    yun_mu2 = pinyin(ch2, style=Style(5), heteronym=True)[0]

    print(sheng_diao1, sheng_mu1, yun_mu1)
    print(sheng_diao2, sheng_mu2, yun_mu2)

    flag1, flag2, flag3 = False, False, False

    for elem in sheng_diao1:
        if elem in sheng_diao2:
            flag1 = True
            break

    for elem in sheng_mu1:
        if elem in sheng_mu2:
            flag2 = True

    for elem in yun_mu1:
        if elem in yun_mu2:
            flag3 = True
    return flag1 or flag2 or flag3


def JudgeCharacter(ch):
    return '\u4e00' <= ch <= '\u9fff'


# class ChineseCharacter:
#     def __init__(self):
#         self.properties = [None] * 25
#         self.dict = {}
#         self.props = []
#
#     def Output(self):
#         for i in range(len(self.props)):
#             print(self.props[i], ":", self.dict[self.props[i]])
#
#
# class CharacterLists:
#     def __init__(self):
#         excel_path = "reading_system/static/character/characters.xlsx"
#         wb = openpyxl.load_workbook(excel_path)
#         ws = wb.active
#
#         lines = []
#         for line in ws:
#             lines.append(line)
#
#         self.characters = [ChineseCharacter()] * len(lines[0])
#
#         cols = len(lines[0])
#         rows = len(lines)
#
#         props = []
#
#         for i in range(rows):
#             props.append(lines[i][1].value)
#
#         print(props)
#
#         cc = 10
#
#         for i in range(2, cc):
#             self.characters[i].props = props
#
#             for j in range(rows):
#                 self.characters[i].properties[j] = lines[j][i].value
#                 self.characters[i].dict[props[j]] = lines[j][i].value
#
#             print(self.characters[i].dict['汉字'])
#
#         for i in range(2, cc):
#             print(self.characters[i].dict['汉字'])
#
#     # 按部件搜索
#     def SearchCharacterByComponent(self, component):
#         pass
#
#     def SearchCharacterByPinyin(self, piny):
#         pass
#
#     def SearchCharacter(self, ch):
#         for elem in self.characters:
#             # print(elem.dict['汉字'])
#             if elem.dict['汉字'] == ch:
#                 return elem.dict
#
#
# def test0():
#     excel_path = "reading_system/static/character/characters.xlsx"
#     wb = openpyxl.load_workbook(excel_path)
#     ws = wb.active
#     cnt = 0
#     lines = []
#     for line in ws:
#         lines.append(line)
#
#
# def test1():
#     chs = CharacterLists()
#     chs.SearchCharacter('天')
#
#
# test1()

#
#


def CheckCharacter(tar, sentence):
    if not JudgeCharacter(sentence[0]):
        return False
    flag1 = IsSimilarChar(tar, sentence[0])
    flag2 = JudgePyInSentence(tar, sentence)
    flag3 = CompareChar(tar, sentence[0]) or CompareChar3(tar, sentence)
    flag = flag3 or (flag1 and flag2)
    return flag


def JudgePyInSentence(tar, sentence):
    res1 = GetPinyin(tar)
    for elem in res1:
        for ch in sentence:
            res2 = GetPinyin(ch)
            if elem in res2:
                return True
    return False


# print(JudgePyInSentence("贪", "谈心"))


# TestExcel()
# print(JudgePyInSentence('披', '拼披萨'))

def CompareChar2(tar, res):
    pyin = ""
    for ch in res:
        if ch.encode('utf-8').isalpha():
            pyin = pyin + ch
        else:
            break
    pyin = pyin.lower()
    sheng_diao = pinyin(tar, style=Style(0), heteronym=True)
    return pyin in sheng_diao[0]


def Check(tar, res):
    substr = res[1:]
    if CompareChar(tar, res[0]) or CompareChar3(tar, res):
        # 第一个字识别正确
        if JudgePyInSentence(tar, substr):
            msg = "正确，识别字正确且组词正确"
        else:
            msg = "正确，识别字正确但组词错误"
    else:
        # 第一个字识别错误
        if JudgePyInSentence(tar, substr):
            if IsSimilarChar(tar, res[0]):
                msg = "正确，识别字错误但在允许范围内，组词正确"
            else:
                msg = "错误，识别字错误但组词正确"
        else:
            msg = "错误，识别字错误且组词错误"
    return msg


def DigitDict():
    file_name = "reading_system/static/character/digits.txt"
    dict = {}
    with open(file_name, 'r', encoding='utf-8') as f:
        while True:
            s = f.readline()
            if not s:
                break
            idx = s.find(':')
            digit = s[0:idx]
            hanzi = s[idx + 1:-1]
            dict[digit] = hanzi
    return dict


def GetErrorMsg(tar, res):
    substr = res[1:]
    flag1 = CompareChar(tar, res[0]) or CompareChar3(tar, res)
    if flag1:
        # 单字阅读正确
        if JudgePyInSentence(tar, substr):
            msg = "正确，单字阅读正确且组词正确"
        else:
            msg = "正确，单字阅读正确但组词错误"
    else:
        # 单字阅读错误
        if JudgePyInSentence(tar, substr):
            if IsSimilarChar(tar, res[0]):
                msg = "正确，单字阅读错误但在允许范围内且组词正确"
            else:
                msg = "错误，单字阅读错误，不在允许范围内但组词正确"
        else:
            msg = "错误，单字阅读错误且组词错误"
    return msg


from reading_system.models import CharacterOfGrade


def TestSQL():
    print(CharacterOfGrade.objects.filter(character='蝠', grade=3).exists())
    row_object = CharacterOfGrade.objects.filter(character='蝠', grade=3).all()
    print(row_object)


# TestSQL()
def CompareChar3(ch1, ch2):
    if len(ch2) != 1:
        return False
    if not JudgeCharacter(ch2):
        return False
    # ch1 和 ch2 的拼音完全匹配
    py1 = GetPinyin(ch1)
    py2 = GetPinyin(ch2)
    for i in py1:
        if i in py2:
            return True
    return False


def TestCorrect(tar, rset):
    flag = False
    for elem in rset:
        # if len(elem) > 1:
        #     elem = elem[0]
        flag = CompareChar3(tar, elem)
        if flag:
            print(elem)
            break
    return flag


def TestChineseCharacter():
    from reading_system.utils import chinesecharacter
    tar = "跷"
    sentence = "壳;翘;俏;撬;窍;鞘;"
    print("目标汉字读音: ", chinesecharacter.GetPinyin(tar))
    for elem in sentence:
        if chinesecharacter.JudgeCharacter(elem):
            print("识别汉字读音: ", chinesecharacter.GetPinyin(elem))
    flag1 = chinesecharacter.JudgePyInSentence(tar, sentence)
    flag2 = chinesecharacter.JudgePyInSentenceStrict(tar, sentence)
    print(flag1, flag2)


def TestExcel():
    from reading_system.utils import chinesecharacter

    # 初始化表格
    cur_time = time.time().__str__()
    excel_path = "reading_system/static/result/result" + cur_time + ".xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Excel'
    ws['A1'] = '目标汉字'
    ws['B1'] = '识别结果'
    ws['C1'] = '判断结果'
    ws['D1'] = '测试时间'
    ws['E1'] = '学生账号'
    ws['F1'] = '测试类型'
    result = models.WavRecognitionResult.objects.all()
    row = 2
    for obj in result:
        tar = obj.target
        res = obj.result

        ws.cell(row, 1).value = tar
        ws.cell(row, 2).value = res

        msg = "识别错误或朗读错误"

        if len(tar) == 1:
            ws.cell(row, 6).value = "识别单字"
            rset = SplitWavResult(res)
            for elem in rset:
                if len(elem) == 1:
                    if CompareSingleCharacter(tar, elem):
                        msg = "朗读正确，" + tar + " 与 " + elem + "读音相同"
                elif len(elem) > 1:
                    if CompareSingleCharacterInSentence(tar, elem):
                        msg = "朗读正确，" + tar + "在词语 " + elem + " 中"
        elif len(tar) < 5:
            ws.cell(row, 6).value = "识别词语"
            ans_len = chinesecharacter.LCS(tar, res)
            ans_str = chinesecharacter.LCS_str(tar, res)
            if len(tar) == ans_len:
                msg = "完全正确"
            elif ans_len > 0:
                msg = "部分正确，朗读正确部分: " + ans_str
        else:
            ws.cell(row, 6).value = "识别句子"
            ans_len = chinesecharacter.LCS(tar, res)
            ans_str = chinesecharacter.LCS_str(tar, res)
            if len(tar) == ans_len:
                msg = "完全正确"
            elif ans_len > 0:
                msg = "部分正确，朗读正确部分: " + ans_str
        ws.cell(row, 3).value = msg
        ws.cell(row, 4).value = obj.exercise_time.__str__()
        ws.cell(row, 5).value = obj.stu
        row += 1
    wb.save(excel_path)
    print("生成分析 Excel 成功")


def DrawWavform():
    import numpy as np
    import matplotlib.pyplot as plt
    from scipy.io import wavfile
    # 读取.wav文件
    sample_rate, data = wavfile.read('reading_system/static/wav/2023629583呼.wav')

    # 将数据转换为浮点数
    data = data.astype(np.float32)

    # 计算时间轴
    time = np.arange(0, len(data)) / sample_rate

    # 绘制波形图
    plt.figure()
    plt.plot(time, data)
    plt.xlabel('Time (s)')
    plt.ylabel('Amplitude')
    plt.title('Waveform of audio.wav')
    plt.show()


def TestFloat():
    sum = 0.0
    a = float("1.0") + 2
    b = "1.0"
    print(a, b)


def TestExercises():
    exercise_list = GenExerciseList(1)
    print(len(exercise_list))
    exercise_list = GenExerciseList(3)
    print(len(exercise_list))


def GenExerciseList(grade):
    from reading_system.utils.chooseList import gen
    import random
    list = []
    while len(list) < 100:
        exercise_list = gen.getChartOriginal()
        if 1 <= grade <= 2:
            # 对应难度 1 2 3 4 5 6
            for level in range(0, 6):
                for ch in exercise_list[level]:
                    elem = {"ch": ch[0], "level": ch[1]}
                    if elem not in list:
                        list.append(elem)
        else:
            for level in range(1, 10):
                for ch in exercise_list[level]:
                    elem = {"ch": ch[0], "level": ch[1]}
                    if elem not in list:
                        list.append(elem)
    random.shuffle(list)
    return list


def TestChineserCharacter():
    from reading_system.utils.chinesecharacter import LCS, LCS_str
    str1 = "你好小明123"
    str2 = "你好世界146"

    print(LCS(str1, str2))
    print(LCS_str(str1, str2))

    print(str1, str2)


def SplitWavResult(wav_result):
    rset = wav_result.split(";")
    return rset


from reading_system.utils.chinesecharacter import characterLists, pyinDict


# 根据拼音字典，比较两个汉字是否是相同读音
def CompareSingleCharacter(tar, src):
    if len(src) != 1 or not JudgeCharacter(src):
        return False
    list1 = pyinDict[tar]
    if src in pyinDict:
        list2 = pyinDict[src]
    else:
        list2 = GetPinyin(src)
    for elem in list1:
        if elem in list2:
            return True
    return False


# 根据拼音字典，查看汉字读音是否在句子中
def CompareSingleCharacterInSentence(tar, src):
    for ch in src:
        if not JudgeCharacter(ch):
            continue
        flag = CompareSingleCharacter(tar, ch)
        if flag:
            return True
    return False


# 判断单个汉字的正确性，rset为候选结果的集合
def JudgeSingleCharacter(tar, rset):
    for elem in rset:
        flag = False
        if len(elem) == 1:
            flag = CompareSingleCharacter(tar, elem)
        elif len(elem) > 1:
            flag = CompareSingleCharacterInSentence(tar, elem)
        if flag:
            return True
    return False


def GetDict(characterLists):
    import re
    pyinDict = {}
    for elem in characterLists.characters:
        # print(elem.dict["汉字"])
        # print(elem.dict["拼音"])
        pyin = elem.dict["拼音"]
        pyinset = re.split(",|;", pyin)
        pyinset = [x for x in pyinset if not x.isdigit()]
        pyinDict[elem.dict["汉字"]] = pyinset
    write_dict_to_file(pyinDict, "reading_system/static/character/pydict.txt")
    return pyinDict


# 测试判别识别结果是否正确
def TestJudge():
    from reading_system.utils.chinesecharacter import pyinDict
    wav_result = "err_msgerr_nosn;打;把;答;拔;波;"
    tar = '拨'
    rset = SplitWavResult(wav_result)
    print(JudgeSingleCharacter(tar, rset))


# 将字典写入文件中
def write_dict_to_file(dictionary, filename):
    import json
    with open(filename, 'w') as file:
        for key, value in dictionary.items():
            file.write(json.dumps({key: value}, ensure_ascii=False) + '\n')


def TestPinyinTable():
    from reading_system.utils import PinYin
    pinyin_table = PinYin.PinYinTable()
    print(pinyin_table.GetVowel("qian"))
    print(pinyin_table.GetConsonant("qian"))
    print(pinyin_table.GetConsonant("ang"))
    from reading_system.utils.chinesecharacter import pyinDict, GetPinyinVec, JudgeBetweenCharacters
    print(JudgeBetweenCharacters("按", "班"))


def TestGetErrorMessage():
    from reading_system.utils.chinesecharacter import GetErrorMessage
    tar = "请"
    res = "晴顷刻"
    print(GetErrorMessage(tar, res))


def TestPinyinVec():
    from reading_system.utils.chinesecharacter import GetPinyinVec, GetPinyin2
    print(GetPinyinVec("世"))
    print(GetPinyinVec("界"))


def TestJudgeSingleCharacterIsTolerable():
    from reading_system.utils.chinesecharacter import JudgeSingleCharacterIsTolerable
    tar = "是"
    rset = ["error", "alkdsa", "界", "田"]
    print(JudgeSingleCharacterIsTolerable(tar, rset))
    # from reading_system.utils.chinesecharacter import GetPinyinVec, GetPinyin2
    # print(GetPinyinVec("世"))
    # print(GetPinyinVec("界"))
    # from reading_system.utils.PinYin import pinyinTable
    # print(pinyinTable.JudgeDifferentSounds("shi", "jie"))
    # c1 = pinyinTable.GetConsonant("shi")
    # c2 = pinyinTable.GetConsonant("jie")
    # print(c1, c2, c1 == c2)
    # v1 = pinyinTable.GetVowel("shi")
    # v2 = pinyinTable.GetVowel("jie")
    # print(v1, v2, v1 == v2)


def TestGetExerciseBank():
    from reading_system.utils.chinesecharacter import GetExerciseBank
    list5 = GetExerciseBank(5)
    list6 = GetExerciseBank(6)
    print("五年级", list5)
    print("六年级", list6)
    print("五年级", len(list5))
    print("六年级", len(list6))


TestGetExerciseBank()
