from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt

from reading_system import models

from io import BytesIO
import openpyxl


def test_upload(request):
    return render(request, "test_upload.html")


@csrf_exempt
def download_excel(request, nid=0):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Excel'
    if nid == 0:
        # 对ws的单个单元格传入数据
        ws['A1'] = '汉字'
        ws['B1'] = '总次数'
        ws['C1'] = '正确次数'
        ws['D1'] = '正确率'
        grade = int(request.POST.get('grade'))
        if grade == 0:
            character_list = models.Character.objects.all()
        else:
            character_list = models.CharacterOfGrade.objects.filter(grade=grade)

        for i in range(0, len(character_list)):
            ws.cell(i + 2, 1).value = character_list[i].character
            ws.cell(i + 2, 2).value = character_list[i].total_time
            ws.cell(i + 2, 3).value = character_list[i].accurate_time
            ws.cell(i + 2, 4).value = character_list[i].accuracy
    elif nid == 1:
        ws['A1'] = '语句'
        ws['B1'] = '出现总次数'
        ws['C1'] = '正确次数'
        ws['D1'] = '正确率'

        grade = int(request.POST.get('grade'))
        if grade == 0:
            exercise_list = models.Exercise.objects.all()
        else:
            exercise_list = models.CharacterOfGrade.objects.filter(grade=grade)

        for i in range(0, len(exercise_list)):
            ws.cell(i + 2, 1).value = exercise_list[i].content
            ws.cell(i + 2, 2).value = exercise_list[i].total
            ws.cell(i + 2, 3).value = exercise_list[i].right
            ws.cell(i + 2, 4).value = exercise_list[i].accuracy

    elif nid == 2:
        ws['A1'] = '学生账号'
        ws['B1'] = '学生姓名'
        ws['C1'] = '年级'
        ws['D1'] = '总字数'
        ws['E1'] = '分数'
        ws['F1'] = '识字量'
        ws['G1'] = '正确率'
        ws['H1'] = '错误汉字'
        ws['I1'] = '测试时间'

        exercise_list = models.ExerciseV1Result.objects.all()

        for i in range(0, len(exercise_list)):
            ws.cell(i + 2, 1).value = exercise_list[i].stu_account
            ws.cell(i + 2, 2).value = exercise_list[i].name
            ws.cell(i + 2, 3).value = exercise_list[i].grade
            ws.cell(i + 2, 4).value = exercise_list[i].total_characters
            ws.cell(i + 2, 5).value = exercise_list[i].score
            ws.cell(i + 2, 6).value = exercise_list[i].literacy
            ws.cell(i + 2, 7).value = exercise_list[i].accuracy_rate
            ws.cell(i + 2, 8).value = exercise_list[i].wrong
            ws.cell(i + 2, 9).value = exercise_list[i].exercise_time.__str__()
    elif nid == 3:
        ws['A1'] = '学生账号'
        ws['B1'] = '学生姓名'
        ws['C1'] = '年级'
        ws['D1'] = '总字数'
        ws['E1'] = '错误数'
        ws['F1'] = '分数'
        ws['G1'] = '测试用时'
        ws['H1'] = '正确率'
        ws['I1'] = '平均阅读速度'
        ws['J1'] = '错误汉字'
        ws['K1'] = '测试时间'

        exercise_list = models.ExerciseV2Result.objects.all()
        for i in range(0, len(exercise_list)):
            ws.cell(i + 2, 1).value = exercise_list[i].stu_account
            ws.cell(i + 2, 2).value = exercise_list[i].name
            ws.cell(i + 2, 3).value = exercise_list[i].grade
            ws.cell(i + 2, 4).value = exercise_list[i].total_characters
            ws.cell(i + 2, 5).value = exercise_list[i].wrong_characters
            ws.cell(i + 2, 6).value = exercise_list[i].score
            ws.cell(i + 2, 7).value = exercise_list[i].test_time
            ws.cell(i + 2, 8).value = exercise_list[i].accuracy_rate
            ws.cell(i + 2, 9).value = exercise_list[i].avg_speed
            ws.cell(i + 2, 10).value = exercise_list[i].wrong
            ws.cell(i + 2, 11).value = exercise_list[i].exercise_time.__str__()
    elif nid == 4:
        ws['A1'] = '学生账号'
        ws['B1'] = '学生姓名'
        ws['C1'] = '年级'
        ws['D1'] = '总字数'
        ws['E1'] = '错误数'
        ws['F1'] = '分数'
        ws['G1'] = '测试用时'
        ws['H1'] = '正确率'
        ws['I1'] = '平均阅读速度'
        ws['J1'] = '测试语句个数'
        ws['K1'] = '判断正确语句个数'
        ws['L1'] = '判断正确率'
        ws['M1'] = '错误汉字'
        ws['N1'] = '测试时间'

        exercise_list = models.ExerciseV3Result.objects.all()
        for i in range(0, len(exercise_list)):
            ws.cell(i + 2, 1).value = exercise_list[i].stu_account
            ws.cell(i + 2, 2).value = exercise_list[i].name
            ws.cell(i + 2, 3).value = exercise_list[i].grade
            ws.cell(i + 2, 4).value = exercise_list[i].total_characters
            ws.cell(i + 2, 5).value = exercise_list[i].wrong_characters
            ws.cell(i + 2, 6).value = exercise_list[i].score
            ws.cell(i + 2, 7).value = exercise_list[i].test_time
            ws.cell(i + 2, 8).value = exercise_list[i].accuracy_rate
            ws.cell(i + 2, 9).value = exercise_list[i].avg_speed
            ws.cell(i + 2, 10).value = exercise_list[i].judge_all
            ws.cell(i + 2, 11).value = exercise_list[i].judge_right
            ws.cell(i + 2, 12).value = exercise_list[i].judge_accuracy
            ws.cell(i + 2, 13).value = exercise_list[i].wrong
            ws.cell(i + 2, 14).value = exercise_list[i].exercise_time.__str__()

    elif nid == 5:
        ws['A1'] = '学生账号'
        ws['B1'] = '学生姓名'
        ws['C1'] = '年级'
        ws['D1'] = '总字数'
        ws['E1'] = '分数'
        ws['F1'] = '识字量'
        ws['G1'] = '正确率'
        ws['H1'] = '错误汉字'
        ws['I1'] = '测试时间'

        name = request.POST.get('name')

        exercise_list = models.ExerciseV1Result.objects.filter(name=name)
        for i in range(0, len(exercise_list)):
            ws.cell(i + 2, 1).value = exercise_list[i].stu_account
            ws.cell(i + 2, 2).value = exercise_list[i].name
            ws.cell(i + 2, 3).value = exercise_list[i].grade
            ws.cell(i + 2, 4).value = exercise_list[i].total_characters
            ws.cell(i + 2, 5).value = exercise_list[i].score
            ws.cell(i + 2, 6).value = exercise_list[i].literacy
            ws.cell(i + 2, 7).value = exercise_list[i].accuracy_rate
            ws.cell(i + 2, 8).value = exercise_list[i].wrong
            ws.cell(i + 2, 9).value = exercise_list[i].exercise_time.__str__()
    elif nid == 7:
        # 某学生的流畅性测试一
        ws['A1'] = '学生账号'
        ws['B1'] = '学生姓名'
        ws['C1'] = '年级'
        ws['D1'] = '总字数'
        ws['E1'] = '错误数'
        ws['F1'] = '分数'
        ws['G1'] = '测试用时'
        ws['H1'] = '正确率'
        ws['I1'] = '平均阅读速度'
        ws['J1'] = '错误汉字'
        ws['K1'] = '测试时间'

        name = request.POST.get('name')
        exercise_list = models.ExerciseV2Result.objects.filter(name=name)

        for i in range(0, len(exercise_list)):
            ws.cell(i + 2, 1).value = exercise_list[i].stu_account
            ws.cell(i + 2, 2).value = exercise_list[i].name
            ws.cell(i + 2, 3).value = exercise_list[i].grade
            ws.cell(i + 2, 4).value = exercise_list[i].total_characters
            ws.cell(i + 2, 5).value = exercise_list[i].wrong_characters
            ws.cell(i + 2, 6).value = exercise_list[i].score
            ws.cell(i + 2, 7).value = exercise_list[i].test_time
            ws.cell(i + 2, 8).value = exercise_list[i].accuracy_rate
            ws.cell(i + 2, 9).value = exercise_list[i].avg_speed
            ws.cell(i + 2, 10).value = exercise_list[i].wrong
            ws.cell(i + 2, 11).value = exercise_list[i].exercise_time.__str__()
    elif nid == 6:
        ws['A1'] = '学生账号'
        ws['B1'] = '学生姓名'
        ws['C1'] = '年级'
        ws['D1'] = '总字数'
        ws['E1'] = '错误数'
        ws['F1'] = '分数'
        ws['G1'] = '测试用时'
        ws['H1'] = '正确率'
        ws['I1'] = '平均阅读速度'
        ws['J1'] = '测试语句个数'
        ws['K1'] = '判断正确语句个数'
        ws['L1'] = '判断正确率'
        ws['M1'] = '错误汉字'
        ws['N1'] = '测试时间'

        name = request.POST.get('name')

        exercise_list = models.ExerciseV3Result.objects.filter(name=name)
        for i in range(0, len(exercise_list)):
            ws.cell(i + 2, 1).value = exercise_list[i].stu_account
            ws.cell(i + 2, 2).value = exercise_list[i].name
            ws.cell(i + 2, 3).value = exercise_list[i].grade
            ws.cell(i + 2, 4).value = exercise_list[i].total_characters
            ws.cell(i + 2, 5).value = exercise_list[i].wrong_characters
            ws.cell(i + 2, 6).value = exercise_list[i].score
            ws.cell(i + 2, 7).value = exercise_list[i].test_time
            ws.cell(i + 2, 8).value = exercise_list[i].accuracy_rate
            ws.cell(i + 2, 9).value = exercise_list[i].avg_speed
            ws.cell(i + 2, 10).value = exercise_list[i].judge_all
            ws.cell(i + 2, 11).value = exercise_list[i].judge_right
            ws.cell(i + 2, 12).value = exercise_list[i].judge_accuracy
            ws.cell(i + 2, 13).value = exercise_list[i].wrong
            ws.cell(i + 2, 14).value = exercise_list[i].exercise_time.__str__()
    elif nid == 8:
        # 所有年级或者某个年级的词语统计结果
        ws['A1'] = '词语'
        ws['B1'] = '出现总次数'
        ws['C1'] = '正确次数'
        ws['D1'] = '正确率'
        grade = int(request.POST.get('grade'))

        if grade == 0:
            exercise_list = models.Phrase.objects.all()
        else:
            exercise_list = models.PhraseOfGrade.objects.filter(grade=grade)
        for i in range(0, len(exercise_list)):
            ws.cell(i + 2, 1).value = exercise_list[i].content
            ws.cell(i + 2, 2).value = exercise_list[i].total
            ws.cell(i + 2, 3).value = exercise_list[i].right
            ws.cell(i + 2, 4).value = exercise_list[i].accuracy

    elif nid == 9:
        from reading_system.utils import chinesecharacter
        # 生成所有识别结果的表格
        ws.append(['目标汉字', '识别结果', '判断结果', '测试时间', '学生账号', '测试类型'])
        # ws['A1'] = '目标汉字'
        # ws['B1'] = '识别结果'
        # ws['C1'] = '判断结果'
        # ws['D1'] = '测试时间'
        # ws['E1'] = '学生账号'
        # ws['F1'] = '测试类型'
        result = models.WavRecognitionResult.objects.all()
        row = 2
        for obj in result:
            tar = obj.target
            res = obj.result

            ws.cell(row, 1).value = tar
            ws.cell(row, 2).value = res

            msg = chinesecharacter.GetErrorMessage(tar, res)

            ws.cell(row, 3).value = msg
            ws.cell(row, 4).value = obj.exercise_time.__str__()
            ws.cell(row, 5).value = obj.stu
            if len(tar) == 1:
                ws.cell(row, 6).value = "识别单字"
            elif len(tar) < 5:
                ws.cell(row, 6).value = "识别词语"
            else:
                ws.cell(row, 6).value = "识别句子"
            row += 1
    elif nid == 10:
        ws['A1'] = '年级'
        ws['B1'] = '学生'
        ws['C1'] = '账号'
        ws['D1'] = '时间'
        ws['E1'] = '识字总数'
        ws['F1'] = '正确率'
        ws['G1'] = '分数'
        ws['H1'] = '错误个数'
        ws['I1'] = '错误汉字'

        exercise_list = models.ExerciseV4Result.objects.all()
        for i in range(0, len(exercise_list)):
            ws.cell(i + 2, 1).value = exercise_list[i].grade
            ws.cell(i + 2, 2).value = exercise_list[i].name
            ws.cell(i + 2, 3).value = exercise_list[i].stu_account
            ws.cell(i + 2, 4).value = exercise_list[i].exercise_time.__str__()
            ws.cell(i + 2, 5).value = exercise_list[i].total_num
            ws.cell(i + 2, 6).value = exercise_list[i].accuracy_rate
            ws.cell(i + 2, 7).value = exercise_list[i].right_num
            ws.cell(i + 2, 8).value = exercise_list[i].wrong_num
            ws.cell(i + 2, 9).value = exercise_list[i].wrong_characters

    # 准备写入到IO中
    output = BytesIO()
    wb.save(output)  # 将Excel文件内容保存到IO中
    output.seek(0)  # 重新定位到开始
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=result.xls'
    return response


@csrf_exempt
def upload_excel(request):
    # 生成一个 Workbook 的实例化对象，wb即代表一个工作簿（一个 Excel 文件）
    wb = openpyxl.Workbook()
    # 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
    ws = wb.active
    # 更改工作表ws的title
    ws.title = 'test_sheet1'
    # 对ws的单个单元格传入数据
    ws['A1'] = '国家'
    ws['B1'] = '首都'
    data = {
        '中国': '北京',
        '韩国': '首尔',
        '日本': '东京',
        '泰国': '曼谷',
        '马来西亚': '吉隆坡',
        '越南': '河内',
        '朝鲜': '平壤',
        '印度': '新德里'
    }
    data_excel = []
    # 将字典中的每对数据（键，值）以列表形式传入data_excel列表
    for each in data:
        data_excel.append([each, data[each]])
    # 将data_excel列表内的内容存入工作表
    for each in data_excel:
        ws.append(each)
    # 注意：上述两个append方法是意义完全不同的两个方法

    # 准备写入到IO中
    output = BytesIO()
    wb.save(output)  # 将Excel文件内容保存到IO中
    output.seek(0)  # 重新定位到开始
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=result.xls'
    return response


def test_voice(request):
    return render(request, "test_voice.html")


def test_fluency(request):
    return render(request, "test_fluency.html")
